import csv
import openpyxl
import pandas as pd
import os
import pdb
from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import  Border, Side, Alignment, Font
import re
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import math
from openpyxl.utils import rows_from_range
from openpyxl.worksheet.cell_range import CellRange



# create equip 
def get_equipments_list(equip_table):
    
# 1: create euiq_dict {
#       "block": "BuildingA',
#       "floor":"Basement 1",
#       "area":"Floor Area",
#       "type":"ASEF",
#       "title":"ASEF-01-01 to ASEF-01-07",
#       "quantity":2,
#       "no_of_points": 5,
#       "smart_points":2,
#       "demarcation": "Base build",
#       "plusSpares":8,
#       "MCC": "MCC-B-CPEF",
#       "ElectricalLoad": FALSE/TRUE
#              }

#### 2: create equi_qty per mcc location {'MMC-A' : {"fcu" : 6, "ahu":9}}
    with open(equip_table,"r") as equipTable:      
        csvreader = csv.reader(equipTable)
        header = next(csvreader)
        equipments_list = [] # type: list 
        equip_qty_per_floor_MCC = []
        block_set = set()
        floor_num_dict = {}
        floor_set = set()
        for row in csvreader:
            if row[3].split("_")[0] == "Others":
           
                equip_dict = {
                    "block": row[0],
                    "floor": row[1],
                    "area":  row[2],
                    "type":f'Others({row[3].split("_",1)[1].replace("_", " ")})',
                    "title":row[4],
                    "quantity":int(row[6]),
                    "no_of_points":int(row[7]),
                    "smart_points":int(row[8]),
                    "demarcation": row[5],
                    "plusSpares":int(row[9]),
                    "MCC":row[10],
                    "Electrical_load":row[11]
                }
            else:
                equip_dict = {
                "block": row[0],
                "floor": row[1],
                "area":  row[2],
                "type":row[3],
                "title":row[4],
                "quantity":int(row[6]),
                "no_of_points":int(row[7]),
                "smart_points":int(row[8]),
                "demarcation": row[5],
                "plusSpares":int(row[9]),
                "MCC":row[10],
                "Electrical_load":row[11]
            }
            equipments_list.append(
                equip_dict
            )
            block_set.add(row[0])
           
        ### only electruical load == true is needed to be filled in electric load table
            if row[11] == "TRUE":
                
                # if row[10] in equip_qty_per_floor_MCC:
                #     equip_qty_per_MCC_dict[row[10]][row[4]] = int(row[6])
                # else:
                #     equip_qty_per_floor_MCC[row[10]]= {row[4]:int(row[6])} 

                equip_qty_per_floor_MCC .append({'MCC': row[10], "floor": row[1], "equip_name": row[4], "QTY": int(row[6])})
        # count floor for point list
            if 'floor' in row[1].lower() or 'level' in row[1].lower():
                floor_set.add(row[1].lower().strip().replace(" ", ""))
                # update when iterate equip
                floor_num_dict[row[0]]  = len(floor_set)
    block_list = list(sorted(block_set))
    #floor_num_dict : the number of the floor in one block {"Block A": 10, "Block B": 11}         [for point summay sheet]
    # block_list: the list of all the blocks in one project ["Block A", "Block B"]                [for iterate in summary sheet]
    # equip_qty_per_floor_MCC: the total quantity of equipment in each floor in one MCC location D1 =[ {'MCC': "MCC01", "Floor":"Floor2", "Equip": "AHU", "QTY": 65},{'MCC': "MCC02", "Floor":"Floor2", "Equip": "AHU", "QTY": 65 }]  
                            #  [for electrical load, change into DF for further calculation]
    return equipments_list,pd.DataFrame(equip_qty_per_floor_MCC),block_list,floor_num_dict

def get_points_list(point_table):  
# point_list
# create euiq_dict {
#       "type":"ASEF",
#       "project_id":12345,
#       "title":"supply temperature",
#       "category": "Standard",
#       "point_type":"Monitor",
#       "point_selected": "Y",
#       "signal_type":"Digital OUT",
#       "cable_type":"BACnet IP",
#       "functional_description":"this is description",
#       
#              }
    point_equip_type_set = set()
    with open(point_table,"r") as pointTable:      
        csvreader = csv.reader(pointTable)
        header = next(csvreader)
        point_list = [] # type: list 
        for row in csvreader:
            
            point_dict = {
                "type":row[0],
                "description":row[1],
                "category": row[3],
                "point_type":row[2],
                "signal_type":row[4],
                "cable_type":row[5],
                "functional_description":row[6],
                "comments":row[7]
            }
            point_list.append(
                point_dict
            )
            point_equip_type_set.add(row[0])
    return point_list, point_equip_type_set


def createPointSummaryDF(equip_table):
###### point summary
    table_row_num =[]
    point_sum_row_num =[]   ### the list of point sum table
    area_count_per_floor_list = {}    # the dict of area count for each block {'Block A': Series(2,3,4,4)}
    tot_point_per_block = {} # the total number of points per block {"Block A": 89} sum of plus spares
    tot_point_per_floor_point = {} # the total number of points per floor {"Block A":series ( 4, 34) } sum of plus spares
    point_total = pd.DataFrame()
    equipments_list,_, block_list, floor_num_dict = get_equipments_list(equip_table)
    for i, block in enumerate(block_list):
        equipments_list_per_block= []
        
        for equip in equipments_list:  
            if equip['block'] == block:
                equipments_list_per_block.append(equip)

        equip_df = pd.DataFrame(equipments_list_per_block)
        equip_df['project_points'] = equip_df['quantity'] * equip_df['no_of_points'] 
        equip_df['totals'] = equip_df['plusSpares'] 
        point_sum =  equip_df[["block","area","demarcation","project_points","smart_points","plusSpares","totals"]]
        point_sum = point_sum.sort_values(["block","area"])
        point_sum = point_sum.groupby(["block","area","demarcation"]).agg({"project_points":'sum',"smart_points":"sum","plusSpares":"sum","totals":"sum"}, inplace = True)
        tot_point_per_block[block] = point_sum.groupby('block').sum()['plusSpares']
        

        point_sum = point_sum.reset_index( drop=False)
        # Rename the columns of each DataFrame to concate
        point_sum_renamed = point_sum.rename(columns={point_sum.columns[i]: 'common_name' for i in range(len(point_sum.columns))})
        # row number of point summary for each block (df) + 1(table1 header) + 1 (table1 header)
        point_sum_row_num.append(point_sum.shape[0] + 2)

###### 1 empty line
 # Create a series to be added between dataframe
        empty_line = pd.Series([" "] * 7)
        title = pd.Series(["Floor Summary","Area","Demarcation", 'Project Points','Smart Points','Plus Spare', 'Totals'])
        #### need to let placeholder keep different otherwise the cell will be merged and later format cannot write in this merged cell
        header_line = pd.Series(["8"] * 7)
        title_placeholder = pd.Series(["*"] * 7)
        

    ###### point breakdown
        equip_df1 = pd.DataFrame(equipments_list)
        equip_df1['project_points'] = equip_df1['quantity'] * equip_df1['no_of_points'] 
        equip_df1['totals'] = equip_df1['plusSpares'] 
        point_bd =  equip_df1[["floor","area","demarcation","project_points","smart_points","plusSpares","totals"]]
        point_bd = point_bd.sort_values(["floor","area","demarcation"])
        point_bd = point_bd.groupby(["floor","area","demarcation"]).agg({"project_points":"sum","smart_points":"sum","plusSpares":"sum","totals":"sum"}, inplace = True)
        point_bd =point_bd.reset_index( drop=False)
        tot_point_per_floor_point[block] = point_bd.groupby(['floor']).sum()['plusSpares']
        point_bd_renamed = point_bd.rename(columns={point_bd.columns[i]: 'common_name' for i in range(len(point_bd.columns))})
        area_count = point_bd.groupby('floor').count()['totals']
        area_count_per_floor_list[block] = area_count
        

  
        if i == len(block_list) - 1:
            # Concatenate the DataFrames
            df_concat = pd.concat([point_sum_renamed,
                               empty_line.set_axis(point_sum_renamed.columns).to_frame().T,  
                               title.set_axis(point_sum_renamed.columns).to_frame().T,
                               point_bd_renamed])
           
        else:
        # Concatenate the DataFrames
            df_concat = pd.concat([point_sum_renamed,
                                empty_line.set_axis(point_sum_renamed.columns).to_frame().T,  
                                title.set_axis(point_sum_renamed.columns).to_frame().T,
                                point_bd_renamed,
                                empty_line.set_axis(point_sum_renamed.columns).to_frame().T,
                                header_line.set_axis(point_sum_renamed.columns).to_frame().T,
                                title_placeholder.set_axis(point_sum_renamed.columns).to_frame().T])
            
        df_concat.columns = ["Summary","Area","Demarcation",'Project Points','Smart Points','Plus Spare', 'Totals']
        df_concat.set_index(['Summary',"Area","Demarcation",'Project Points'], inplace= True)
        # block breakdown header(1) + point summary title(1) + point summary(only data no title) + emptyline(1) + title(1) + point breakdown row (only df no title)
        
        table_row_num.append(point_sum_renamed.shape[0] + point_bd_renamed.shape[0] + 4)
        point_total =  pd.concat([point_total,df_concat])


    #point_sum_row_num
    return point_total, point_sum_row_num, area_count_per_floor_list,block_list, table_row_num,tot_point_per_block,tot_point_per_floor_point

def createEquipSummaryDF(equip_table):
    #record the total row num of each table(block)
    table_row_num =[]
    equip_sum_row_num =[]
    equip_total = pd.DataFrame()
    equip_count_per_floor_list = {}
    tot_point_per_block = {}
    tot_point_per_floor_equip = {}
######  equip smmmary  
    equipments_list, _ , block_list, _= get_equipments_list(equip_table)
    
######FILTER ACCORDING TO BLOCK
    for i, block in enumerate(block_list):
        equipments_list_per_block= []
        
        for equip in equipments_list:  
            if equip['block'] == block:
                equipments_list_per_block.append(equip)
         

        equip_df = pd.DataFrame(equipments_list_per_block)
        equip_df['totals'] = equip_df['plusSpares'] 
        equip_df['sub_totals'] = equip_df['quantity'] * equip_df['plusSpares'] 
        equip_sum =  equip_df[["block","type","demarcation","quantity","no_of_points","smart_points","plusSpares","sub_totals","totals"]]
        equip_sum = equip_sum.sort_values(["block","type","demarcation"])  
        equip_sum = equip_sum.groupby(["block","type","demarcation"]).agg({"quantity":'sum',"no_of_points":"mean", "smart_points":"sum","plusSpares":"sum","sub_totals":"sum", "totals":"sum"}, inplace = True)
        tot_point_per_block[block] = equip_sum.groupby('block').sum()['sub_totals']
        equip_sum = equip_sum.reset_index(drop=False)
        # Rename the columns of each DataFrame to concate
        equip_sum_renamed = equip_sum.rename(columns={equip_sum.columns[i]: 'common_name' for i in range(len(equip_sum.columns))})

        equip_sum_row_num.append(equip_sum.shape[0] + 2)
        
    ###### 1 empty line
    # Create a series to be added between dataframe
        empty_line = pd.Series([" "] * 9)
        title = pd.Series(["Floor Summary","Equipment Type","Demarcation", 'Total Quantities','Unit Points','Smart Points','Plus Spares', 'Sub-Totals','Totals'])
        #### otherwise the cell will be merged and later format cannot write in this merged cell
        header_line = pd.Series(["8"] * 9)
        title_placeholder = pd.Series(["*"] * 9)

    # equip breakdown 
        equip_df1 = pd.DataFrame(equipments_list_per_block)
        equip_df1['totals'] = equip_df1['plusSpares'] 
        equip_df1['sub_totals'] = equip_df1['quantity'] * equip_df1['plusSpares'] 
        equip_bd =  equip_df1[["floor","type","demarcation","quantity","no_of_points","smart_points","plusSpares","sub_totals","totals"]]
        equip_bd = equip_bd.sort_values(["floor","type","demarcation"])
        equip_bd = equip_bd.groupby(["floor","type","demarcation"]).agg({"quantity":'sum',"no_of_points":"mean", "smart_points":"sum","plusSpares":"sum","sub_totals":"sum", "totals":"sum"}, inplace = True)
        equip_bd = equip_bd.reset_index( drop=False)
        equip_count = equip_bd.groupby('floor').count()['totals']
        equip_count_per_floor_list[block] = equip_count
        
        tot_point_per_floor_equip[block] = equip_bd.groupby('floor').sum()['sub_totals']
        equip_bd_renamed = equip_bd.rename(columns={equip_bd.columns[i]: 'common_name' for i in range(len(equip_bd.columns))})

        if i == len(block_list) - 1:
            # last table no need to add placeholder
          
            equip_concat = pd.concat([equip_sum_renamed,
                                      empty_line.set_axis(equip_sum_renamed.columns).to_frame().T,  
                                      title.set_axis(equip_sum_renamed.columns).to_frame().T,
                                      equip_bd_renamed]) 
        # Concatenate the DataFrames
        else:
            equip_concat = pd.concat([equip_sum_renamed,
                                      empty_line.set_axis(equip_sum_renamed.columns).to_frame().T, 
                                      title.set_axis(equip_sum_renamed.columns).to_frame().T,
                                      equip_bd_renamed,
                                      empty_line.set_axis(equip_sum_renamed.columns).to_frame().T,
                                      header_line.set_axis(equip_sum_renamed.columns).to_frame().T,
                                      title_placeholder.set_axis(equip_sum_renamed.columns).to_frame().T])
            
        # header (breakdown)(1) +  equip summary title(1) + point summary(only data no title) + emptyline(1) + title(1) + point breakdown row (only df no title)
        table_row_num.append(equip_sum_renamed.shape[0] + equip_bd_renamed.shape[0] + 4)
        equip_total =  pd.concat([equip_total,equip_concat])


    equip_total.columns = ["Floor Summary","Equipment Type","Demarcation",'Total Quantities','Unit Points','Smart Points', 'Plus Spares', 'Sub-Totals' ,'Totals']
    equip_total.set_index(['Floor Summary',"Equipment Type","Demarcation",'Total Quantities'], inplace= True)
    # total row num  of equip sum(table 1) 1 header + 1 title + df
    
    return equip_total,equip_sum_row_num,equip_count_per_floor_list,block_list, table_row_num,tot_point_per_block, tot_point_per_floor_equip

def create_point_sum_format(writer,sheet_name,point_sum_row_num,point_sum_df,area_count_per_floor_list,block_list,table_row_num,tot_point_per_block,tot_point_per_floor_point):
    
    # point_sum_iloc (x, 3)
    # point_sum_df.index.values[0][3] 
    workbook  = writer.book
    worksheet = writer.sheets[sheet_name]
 
     #set whole sheet format
    workbook.formats[0].set_font_name('Cambria')
    workbook.formats[0].set_font_size(8)
    workbook.formats[0].set_align('center')
    workbook.formats[0].set_align('vcenter')

     ## set width
    worksheet.set_column('B:B', 29)
    worksheet.set_column('C:C', 20)
    worksheet.set_column('D:D', 20) 
    worksheet.set_column('E:E', 15)
    worksheet.set_column('F:F', 15)
    worksheet.set_column('G:G', 15)
    worksheet.set_column('H:H', 10)

    offset = 0
    
    # set cannot access by index, so convert it into list to get index
    
    for i, block in enumerate(block_list):
        ######### HEADER STYLE
        # Add Header format
        format_header = workbook.add_format({'bg_color': '#094121','bold': True,'font_color': 'white','font_size':16,'align': 'center','font_name':'Cambria','valign': 'vcenter','bottom':5,'top':5, 'left':5, 'right':5})
    
        # set Header place
        title =  f"{block} - Floor Breakdown"
        worksheet.merge_range(f'B{2+offset}:H{2+offset}',title,format_header)
        worksheet.set_row(1 + offset, 45)

        ##### set 1st table title
        
        border_format = workbook.add_format({'bold': True,'bottom':5,'top':5, 'left':5, 'right':5,'font_name':'Cambria','font_size':10,'valign': 'vcenter','bold': True,'align': 'center'})
        
        worksheet.set_row(2 + offset, 33)
        worksheet.write(f'B{3+offset}',"Building Summary",border_format)
        worksheet.write(f'C{3+offset}',"Locations",border_format)
        worksheet.write(f'D{3+offset}',"Demarcation",border_format)
        worksheet.write(f'E{3+offset}',"Project points",border_format)
        worksheet.write(f'F{3+offset}',"Smart Points",border_format)
        worksheet.write(f'G{3+offset}',"Plus Spares",border_format)
        worksheet.write(f'H{3+offset}',"Totals",border_format)


        ##### summary total merge 
        ## 3 is the 1 header 2 title 1 jumpback to last row
        
        worksheet.merge_range(f'H{4 + offset}:H{4 + offset + point_sum_row_num[i] - 3}', tot_point_per_block[block])



        ######## 2nd TITLE STYLE
        # title border formate
        # same num in excel
      
        border_format = workbook.add_format({'bg_color': '#094121','bold': True,'font_color': 'white','bottom': 5,'top':5, 'left':5, 'right':5,'font_name':'Cambria','font_size':10,'valign': 'vcenter','bold': True,'align': 'center'})
        # set_row start from 0---> row num - 1
        # empty 1(begin of table) + empty line (end of table) + 1 new line 
        point_bd_row = 3 + offset + point_sum_row_num[i]
        worksheet.set_row(point_bd_row -1, 33)
        worksheet.write(f'B{point_bd_row}',"Floor Summary",border_format)
        worksheet.write(f'C{point_bd_row}',"Area",border_format)
        worksheet.write(f'D{point_bd_row}',"Demarcation",border_format)
        worksheet.write(f'E{point_bd_row}',"Project Points",border_format)
        worksheet.write(f'F{point_bd_row}',"Smart Points",border_format)
        worksheet.write(f'G{point_bd_row}',"Plus Spare",border_format)
        worksheet.write(f'H{point_bd_row}',"Totals",border_format)
        print(os.getcwd())
        worksheet.insert_image("B2", "logo.png",{"x_offset": 30, "y_offset": 10})

    #####  MERGE total
        #1 empty at beginig of table A , 1 empty row at end of table A + 1 2nd table title + 1 start from next one
        
        start = offset + point_sum_row_num[i] + 4
        for num , count in enumerate(area_count_per_floor_list[block]):
            
            worksheet.merge_range(f'H{start}:H{start + count - 1}', tot_point_per_floor_point[block].iloc[num])
            start += count 

        offset += table_row_num[i]  + 1
       
        
        
 
def create_equip_sum_format(writer,sheet_name,equip_sum_row_num,equip_sum_df,equip_count_per_floor_list,block_list,table_row_num,tot_point_per_block_equip, tot_point_per_floor_equip):
    ##equip_sum_row_num: [75,5]
    ##equip_sum_df.index(75, 4)
    workbook  = writer.book
    worksheet = writer.sheets[sheet_name]
     #set whole sheet format
    workbook.formats[0].set_font_name('Cambria')
    workbook.formats[0].set_font_size(8)
    workbook.formats[0].set_align('center')
    workbook.formats[0].set_align('vcenter')

     ## set width
    worksheet.set_column('B:B', 29)
    worksheet.set_column('C:C', 29)
    worksheet.set_column('D:D', 20) 
    worksheet.set_column('E:E', 18)
    worksheet.set_column('F:F', 15)
    worksheet.set_column('G:G', 15)
    worksheet.set_column('H:H', 15)
    worksheet.set_column('I:I', 15)
    worksheet.set_column('J:J', 10)

    offset = 0
    
    # set cannot access by index, so convert it into list to get index
    

    for i, block in enumerate(block_list):

        ######### HEADER STYLE
        # Add Header format
        format_header = workbook.add_format({'bg_color': '#094121','bold': True,'font_color': 'white','font_size':16,'align': 'center','font_name':'Cambria','valign': 'vcenter','bottom': 5,'top':5, 'left':5, 'right':5})
    
        # set Header place
        title =  f"{block} - Equipment Breakdown"
    
        worksheet.merge_range(f'B{2 + offset}:J{2 + offset}',title,format_header)
        
        worksheet.set_row(1 + offset, 45) 

        ####### 1st table title style
        border_format = workbook.add_format({'bold': True,'bottom': 5,'top':5, 'left':5, 'right':5,'font_name':'Cambria','font_size':10,'valign': 'vcenter','bold': True,'align': 'center'})
        ###setrow start from 0
        worksheet.set_row(2 + offset, 33)
        worksheet.write(f'B{3 + offset}',"Building Summary",border_format)
        worksheet.write(f'C{3 + offset}',"Equipment Type",border_format)
        worksheet.write(f'D{3 + offset}',"Demarcation",border_format)
        worksheet.write(f'E{3 + offset}',"Total Quantities",border_format)
        worksheet.write(f'F{3 + offset}',"Unit Points",border_format)
        worksheet.write(f'G{3 + offset}',"Smart Points",border_format)
        worksheet.write(f'H{3 + offset}',"Plus Spares",border_format)
        worksheet.write(f'I{3 + offset}',"Sub-Totals",border_format)
        worksheet.write(f'J{3 + offset}',"Totals",border_format)

        
        ### summary total merge 
        # 3 is the 1 header 2 title 1 jumpback to last row
        worksheet.merge_range(f'J{4 + offset}:J{4 + offset + equip_sum_row_num[i] - 3 }', tot_point_per_block_equip[block])


        ####### 2nd TITLE STYLE
        #title border formate
        border_format = workbook.add_format({'bg_color': '#094121','bold': True,'font_color': 'white','bottom': 5,'top':5, 'left':5, 'right':5,'font_name':'Cambria','font_size':10,'valign': 'vcenter','bold': True,'align': 'center'})
        # empty 1(begin of table) + empty line (end of table) + 1 new line 
        equip_bd_row = 3 + offset + equip_sum_row_num[i]
        worksheet.set_row(equip_bd_row - 1, 33)
        worksheet.write(f'B{equip_bd_row}',"Floor Summary",border_format)
        worksheet.write(f'C{equip_bd_row}',"Equipment Type",border_format)
        worksheet.write(f'D{equip_bd_row}',"Demarcation",border_format)
        worksheet.write(f'E{equip_bd_row}',"Total Quantities",border_format)
        worksheet.write(f'F{equip_bd_row}',"Unit Points",border_format)
        worksheet.write(f'G{equip_bd_row}',"Smart Points",border_format)
        worksheet.write(f'H{equip_bd_row}',"Plus Spares",border_format)
        worksheet.write(f'I{equip_bd_row}',"Sub-Totals",border_format)
        worksheet.write(f'J{equip_bd_row}',"Totals",border_format)
        worksheet.insert_image("B2", "logo.png",{"x_offset": 30, "y_offset": 10})

  
        #######  MERGE total
        # 1 empty at beginig of table A , 1 empty row at end of table A + 1 2nd table title + start from next one
       
        start = offset + equip_sum_row_num[i] + 4
        for num , count in enumerate( equip_count_per_floor_list[block]):
            #worksheet.merge_range(f'J{start}:J{ start + count - 1}', tot_point_per_floor_equip[block].iloc[c])
            
            worksheet.merge_range(f'J{start}:J{ start + count - 1}',tot_point_per_floor_equip[block].iloc[num])
            start += count 
           




        offset += table_row_num[i] + 1
        
    
        
def createPointsDF(group,point_equip_type_set, point_list ):       
    filter_equip = group[["Equipment","Title","Floor","Demarcation","Quantity"]]
    filter_equip.sort_values(["Equipment","Title","Floor"])
    # get qty info based on groupby character
    qty_info = filter_equip.groupby(["Equipment","Title","Floor","Demarcation"]).sum()
    point_records_list=[]
    for pinfo , pty in qty_info.iterrows():
        #if equip info can br fount in point list (try catch?)
        if pinfo[0] in point_equip_type_set:
            #find pointdescroption
            
            for p in  point_list:
                # find point list according to the key
                if p['type'] == pinfo[0] :
                    if pinfo[3] == "Landlord":
                        # Other_core -> Other(Core)
                        if pinfo[0].split("_")[0] == "Others":
                            point_record = {
                                    "Equipment": f"Others({pinfo[0].split('_')[1].replace('_', ' ')})",
                                    "Equipment Tag": f"Others({pinfo[1].split('_')[1].replace('_', ' ')})",
                                    "LL-QTY": pty.Quantity,
                                    "FO-QTY":0,
                                    "Floor": pinfo[2],
                                    "Point Description": p["description"],
                                    "Points Type":p["point_type"],
                                    "Signal Type":p["signal_type"],
                                    "Cable Type": p["cable_type"],
                                    "Functional Description":p["functional_description"],
                                    "Comments":""
                                    
                                }
                        else:
                            point_record = {
                                    "Equipment": pinfo[0],
                                    "Equipment Tag": pinfo[1],
                                    "LL-QTY": pty.Quantity,
                                    "FO-QTY":0,
                                    "Floor": pinfo[2],
                                    "Point Description": p["description"],
                                    "Points Type":p["point_type"],
                                    "Signal Type":p["signal_type"],
                                    "Cable Type": p["cable_type"],
                                    "Functional Description":p["functional_description"],
                                    "Comments":""
                                    
                                }
                
                    else:
                        if pinfo[0].split("_")[0] == "Others":
                            point_record = {
                                        "Equipment":  f"Others({pinfo[0].split('_')[1]})",
                                        "Equipment Tag": f"Others({pinfo[1].split('_')[1]})",
                                        "LL-QTY": 0,
                                        "FO-QTY":pty.Quantity,
                                        "Floor": pinfo[2],
                                        "Point Description": p["description"],
                                        "Points Type":p["point_type"],
                                        "Signal Type":p["signal_type"],
                                        "Cable Type": p["cable_type"],
                                        "Functional Description":p["functional_description"],
                                        "Comments":""
                                    }   
                        else:
                            point_record = {
                                    "Equipment": pinfo[0],
                                    "Equipment Tag": pinfo[1],
                                    "LL-QTY": 0,
                                    "FO-QTY":pty.Quantity,
                                    "Floor": pinfo[2],
                                    "Point Description": p["description"],
                                    "Points Type":p["point_type"],
                                    "Signal Type":p["signal_type"],
                                    "Cable Type": p["cable_type"],
                                    "Functional Description":p["functional_description"],
                                    "Comments":""
                                }           
                    point_records_list.append(point_record)
    
    point_record_df = pd.DataFrame( point_records_list)
    point_record_df.set_index(["Equipment","Equipment Tag","LL-QTY","FO-QTY","Floor","Point Description"], inplace = True)
    return point_record_df
        #create_report_per_MCC( point_record_df,MCCname[0],writer)


def createPointsDF2(group,point_equip_type_set, point_list ):  
    group.sort_values(["Equipment","Title","Floor"])
    point_total = pd.DataFrame()
    row_number_list = []
    
    floor_name_list = []
    
    total_BMS_point_list_per_floor=[]
    for floor_name , group_per_floor in group.groupby(['Floor']):
        floor_name_list.append(floor_name[0])
        point_record_per_floor = []
        
        
        filter_equip = group_per_floor[["Equipment","Title","Floor","Demarcation","Quantity"]]
        filter_equip.sort_values(["Equipment","Title","Floor"])
        # get qty info based on groupby character
        qty_info = filter_equip.groupby(["Equipment","Title","Floor","Demarcation"]).sum()
        
        total_point_per_floor = 0
        empty_line = pd.Series([" "] * 11)
        total_placeholder = pd.Series(["x"] * 11)
        header_placeholder = pd.Series(["h"] * 11)
        title_placeholder = pd.Series(["t"] * 11)
        for pinfo , pty in qty_info.iterrows():
            
            #qty.info: pinfo[0]: FCU   pty 65
            point_records_list_per_equip=[]
            
            #if equip info can be found in point list (DB)
            if pinfo[0] in point_equip_type_set:
                
                #find pointdescroption     
                for p in  point_list:
                    # find point list according to the key p
                    if p['type'] == pinfo[0] :
                        # count += 1
                        
                         #to fill the qty into LL or FO
                        if pinfo[3] == "Landlord":
                            # Other_core -> Other(Core)
                            if pinfo[0].split("_")[0] == "Others":
                                point_record = {
                                        "Equipment": f"Others({pinfo[0].split('_')[1].replace('_', ' ')})",
                                        "Equipment Tag": f"Others({pinfo[1].split('_')[1].replace('_', ' ')})",
                                        "LL-QTY": pty.Quantity,
                                        "FO-QTY":0,
                                        "Floor": pinfo[2],
                                        "Point Description": p["description"],
                                        "Points Type":p["point_type"],
                                        "Signal Type":p["signal_type"],
                                        "Cable Type": p["cable_type"],
                                        "Functional Description":p["functional_description"],
                                        "Comments":""
                                        
                                    }
                            else:
                                point_record = {
                                        "Equipment": pinfo[0],
                                        "Equipment Tag": pinfo[1],
                                        "LL-QTY": pty.Quantity,
                                        "FO-QTY":0,
                                        "Floor": pinfo[2],
                                        "Point Description": p["description"],
                                        "Points Type":p["point_type"],
                                        "Signal Type":p["signal_type"],
                                        "Cable Type": p["cable_type"],
                                        "Functional Description":p["functional_description"],
                                        "Comments":""
                                        
                                    }
                
                        else:
                            if pinfo[0].split("_")[0] == "Others":
                                point_record = {
                                            "Equipment":  f"Others({pinfo[0].split('_')[1]})",
                                            "Equipment Tag": f"Others({pinfo[1].split('_')[1]})",
                                            "LL-QTY": 0,
                                            "FO-QTY":pty.Quantity,
                                            "Floor": pinfo[2],
                                            "Point Description": p["description"],
                                            "Points Type":p["point_type"],
                                            "Signal Type":p["signal_type"],
                                            "Cable Type": p["cable_type"],
                                            "Functional Description":p["functional_description"],
                                            "Comments":""
                                        }   
                            else:
                                
                                point_record = {
                                        "Equipment": pinfo[0],
                                        "Equipment Tag": pinfo[1],
                                        "LL-QTY": 0,
                                        "FO-QTY":pty.Quantity,
                                        "Floor": pinfo[2],
                                        "Point Description": p["description"],
                                        "Points Type":p["point_type"],
                                        "Signal Type":p["signal_type"],
                                        "Cable Type": p["cable_type"],
                                        "Functional Description":p["functional_description"],
                                        "Comments":""
                                    }
                        point_records_list_per_equip.append(point_record)
                        point_record_per_floor.append(point_record)
            
                
            else:
                print("This equipment cannot be found in point list table")
            
            total_point_per_equip = pty.Quantity * len(point_records_list_per_equip)
            
            total_point_per_floor += total_point_per_equip

        
        point_record_per_floor = pd.DataFrame( point_record_per_floor)  
        row_number_list.append(point_record_per_floor.shape[0] + 2)
        total_BMS_point_list_per_floor.append(total_point_per_floor)
           
            

        if point_total.empty:  
            
            point_total = pd.concat([point_total, 
                                    point_record_per_floor])
        else:
           
            point_total = pd.concat([point_total,
                                     empty_line.set_axis(point_record_per_floor.columns).to_frame().T,
                                    total_placeholder.set_axis(point_record_per_floor.columns).to_frame().T,
                                    empty_line.set_axis(point_record_per_floor.columns).to_frame().T,
                                    header_placeholder.set_axis(point_record_per_floor.columns).to_frame().T,
                                    title_placeholder.set_axis(point_record_per_floor.columns).to_frame().T,
                                    point_record_per_floor])
    
    
        
    point_total.set_index(["Equipment","Equipment Tag","LL-QTY","FO-QTY","Floor","Point Description"], inplace = True)
    # point_total: point detail for all floor (total dataframe)
    # floor_name list: the list of floor name for header
    # row_number_list: total point (row num) per floor
    # total row number per floor(Header + title + content) NO EMPTY ROW
    # total_BMS_point_list_per_floor : total BMS points per floor(for total value)
    return point_total, row_number_list, floor_name_list,total_BMS_point_list_per_floor



def create_point_list_sheet_format(tot_df,MCCname,row_number_list, floor_name_list,total_BMS_point_list_per_floor,spare_points,writer):
    # Convert the merged dataframe to an XlsxWriter Excel object.
   
    sheet_name = f"{MCCname} Point List"
    header_start_row = 1
    tot_df.to_excel(writer, sheet_name= sheet_name,startrow = 2,startcol = 1)

    # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets[sheet_name]

    #set whole sheet format
    workbook.formats[0].set_font_name('Cambria')
    workbook.formats[0].set_font_size(8)


    border_format = workbook.add_format({'bottom': 2,'top':4, 'left':4, 'right':4,'font_name':'Cambria','font_size':8})


    ## set width
    worksheet.set_column('B:B', 30)
    worksheet.set_column('C:C', 30)
    worksheet.set_column('D:D', 10) 
    worksheet.set_column('E:E', 10)
    worksheet.set_column('F:F', 20)
    worksheet.set_column('G:G', 35)
    worksheet.set_column('H:H', 30)
    worksheet.set_column('I:I', 30)
    worksheet.set_column('J:J', 30)
    worksheet.set_column('K:K', 120)
    worksheet.set_column('L:L', 30)

    offset = 0
    
    for count, floor_name in enumerate(floor_name_list):

        ######### HEADER STYLE
        # Add Header format
        format_header = workbook.add_format({'bg_color': '#094121','bold': True,'font_color': 'white','font_size':16,'align': 'center','font_name':'Cambria','valign': 'vcenter','bottom': 2,'top':2, 'left':2, 'right':2})
    
        # set Header place
        title =  f"{MCCname} - {floor_name}"
        
        worksheet.merge_range(f'B{2 + offset} :L{2 + offset}',title,format_header)
        worksheet.set_row(header_start_row + offset, 45)
        
            ######## TITLE STYLE
        # title border formate
        border_format = workbook.add_format({'bottom': 2,'top':2, 'left':2, 'right':2,'font_name':'Cambria','font_size':11,'valign': 'vcenter','bold': True,'align': 'center'})
        worksheet.set_row(2 + offset, 33)
        worksheet.write(f'B{3 + offset}',"Equipment",border_format)
        worksheet.write(f'C{3 + offset}',"Equipment Tag",border_format)
        worksheet.write(f'D{3 + offset}',"LL-QTY",border_format)
        worksheet.write(f'E{3 + offset}',"FO-QTY",border_format)
        worksheet.write(f'F{3 + offset}',"Floor",border_format)
        worksheet.write(f'G{3 + offset}',"Point Description",border_format)
        worksheet.write(f'H{3 + offset}',"Points Type",border_format)
        worksheet.write(f'I{3 + offset}',"Signal Type",border_format)
        worksheet.write(f'J{3 + offset}',"Cable Type",border_format)
        worksheet.write(f'K{3 + offset}',"Functional Description",border_format)
        worksheet.write(f'L{3 + offset}',"Comments",border_format)



    ######## TABLE BORDER STYLE
    
    
        ######## CELL STYLE dashed and center (equip col to  floor col)
        cell1_format = workbook.add_format({'bottom': 4,'top':4, 'left':4, 'right':4,'font_name':'Cambria','font_size':8,'align': 'center','valign': 'vcenter'})
        left_border_format = workbook.add_format({'font_name':'Cambria','valign': 'vcenter','align': 'center','font_size':8,'right':4,'left':2,'bottom': 4,'top':4})
        right_border_format = workbook.add_format({'font_name':'Cambria','valign': 'vcenter','align': 'center','font_size':8,'right':2,'left':4,'bottom': 4,'top':4})
        bottom_border_format = workbook.add_format({'font_name':'Cambria','valign': 'vcenter','align': 'center','font_size':28,'right':4,'left':4,'bottom': 2,'top':4})
        description_cell__format = workbook.add_format({'font_name':'Cambria','align': 'left','font_size':8,'right':4,'left':2,'bottom': 4,'top':4})
        # all bold stype
        all_bold_format = workbook.add_format({'font_name':'Cambria','align': 'center','valign': 'vcenter','font_size':10,'right':2,'left':2,'bottom': 2,'top':2,'bold': True,})

        df_tot_row = row_number_list[count] - 2
        for row in range(df_tot_row):
         
            # row_num_list contain title and header, so - 2
            for col in range(5):
               
                if col == 0:            
                    # left border
                    worksheet.write(row + 3 + offset, col + 1, tot_df.index.values[row + offset][col], left_border_format)
                elif col == 4:
                    # right border
                    worksheet.write(row + 3 + offset, col + 1, tot_df.index.values[row + offset][col], right_border_format)          
                else:
                    worksheet.write(row + 3 + offset, col + 1, tot_df.index.values[row + offset][col], cell1_format)

                # last row
                if row == df_tot_row - 1 :
                    # bottom border
                
                    worksheet.write(row + 3 + offset, col + 1,tot_df.index.values[row + offset][col], bottom_border_format)

            #description have diff format
            worksheet.write(row + 3 + offset, 6, tot_df.index.values[row + offset][5], description_cell__format)


        description_cell__button_format = workbook.add_format({'font_name':'Cambria','align': 'left','font_size':8,'right':2,'left':2,'bottom': 2,'top':4})
        worksheet.write(df_tot_row + 2 + offset, 6, tot_df.index.values[df_tot_row - 1  + offset][5], description_cell__button_format)
        # left bottom corner
        left_corner_format = workbook.add_format({'font_name':'Cambria','align': 'left','font_size':8,'right':4,'left':2,'bottom': 2,'top':4})
        worksheet.write(df_tot_row + 2 + offset, 1, tot_df.index.values[df_tot_row - 1 + offset][1], left_corner_format)
            




            ###### CELL STYLE dashed and left align (pd col to  comment col)
        cell2_format = workbook.add_format({'bottom': 4,'top':4, 'left':4, 'right':4,'font_name':'Cambria','font_size':8,'align': 'left'})
        left_border_alignleft_format = workbook.add_format({'font_name':'Cambria','valign': 'vcenter','align': 'left','font_size':8,'right':4,'left':2,'bottom': 4,'top':4})
        right_border__aligenleft_format = workbook.add_format({'font_name':'Cambria','valign': 'vcenter','align': 'left','font_size':8,'right':2,'left':4,'bottom': 4,'top':4})
        for row in range(df_tot_row):
            for col in range(tot_df.shape[1]):
                if col == tot_df.shape[1] - 1:
                    # right border
                    worksheet.write(row + 3 + offset, col + 7, tot_df.iloc[row + offset ,col], right_border__aligenleft_format)         
                else:
                    worksheet.write(row + 3 + offset, col + 7, tot_df.iloc[row + offset,col], cell2_format)
                if row == df_tot_row - 1:
                    # bottom border
                    bottom_border_left_format = workbook.add_format({'font_name':'Cambria','valign': 'vcenter','align': 'left','font_size':8,'right':4,'left':4,'bottom': 2,'top':4})
                    worksheet.write(row + 3 + offset, col + 7, tot_df.iloc[row + offset,col], bottom_border_left_format)
        # right corner
        right_corner_format = workbook.add_format({'font_name':'Cambria','align': 'left','font_size':8,'right':2,'left':4,'bottom': 2,'top':4})
        worksheet.write(df_tot_row + 2 + offset, 11, tot_df.iloc[df_tot_row - 1 + offset, 4], right_corner_format)



        # write the total BMS point
        worksheet.write(df_tot_row + 4 + offset, 2 , total_BMS_point_list_per_floor[count], all_bold_format)
        worksheet.write(df_tot_row + 4 + offset, 7 , math.ceil(total_BMS_point_list_per_floor[count] * (1 + spare_points)), all_bold_format)
        worksheet.write(df_tot_row + 4 + offset, 1 , "Total BMS Points", all_bold_format)
        # merge range start from 0. so need to add 4 + 1
        worksheet.merge_range(f'E{df_tot_row + 5 + offset} :G{df_tot_row + 5 + offset}',f"Total BMS Points (+ {format(spare_points, '.2%')} Spare Capacity)",all_bold_format)


        # + 1 empty row + 1 total row + 1 emoty row
        offset += row_number_list[count] + 3

    # write the total BMS point per MCC (total)
    worksheet.write(tot_df.shape[0] + 7 , 2 , sum(total_BMS_point_list_per_floor), all_bold_format)
    worksheet.write(tot_df.shape[0] + 7 , 7 , sum(total_BMS_point_list_per_floor) * (1 + spare_points), all_bold_format)
    worksheet.write(tot_df.shape[0] + 7 , 1 , f" {MCCname} :Total BMS Points ", all_bold_format)
    # merge range start from 0. so need to add + 1
    worksheet.merge_range(f'E{tot_df.shape[0] + 8 } :G{tot_df.shape[0] + 8}',f"{MCCname}: Total BMS Points (+ {format(spare_points, '.2%')} Spare Capacity)",all_bold_format)
    worksheet.insert_image("B2", "logo.png",{"x_offset": 30, "y_offset": 10})

    

def createReportExcel(point_file, equip_file, project_file):
    project_name = point_file.split("-")[1]
    report_date = point_file.split("-")[2] + "-" + point_file.split("-")[3] +"-" + point_file.split("-")[4].split(".")[0]
    report_name = f'\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\{project_name} - Points Schedule - {report_date}.xlsx'
    file_path = os.path.abspath(report_name)
   
    equip_sum,equip_sum_row_num, equip_sum_count_list,block_list,equip_table_row_num, tot_point_per_block_equip, tot_point_per_floor_equip = createEquipSummaryDF(equip_file)
    point_sum,point_sum_row_num ,point_sum_count_list,block_list,point_table_row_num, tot_point_per_block, tot_point_per_floor_point= createPointSummaryDF(equip_file)

    spare_point = pd.read_csv(project_file)
    spare_points = spare_point['SpareBMSPoints'][0]
    with pd.ExcelWriter(file_path, engine = 'xlsxwriter') as excel_writer:
       # project_sum.to_excel(excel_writer, sheet_name='Points Summary', startrow= 2, startcol=1)
        point_sum.to_excel(excel_writer, sheet_name='Points Summary', startrow= 2, startcol=1)
        create_point_sum_format(excel_writer,'Points Summary',point_sum_row_num, point_sum,point_sum_count_list,block_list,point_table_row_num,tot_point_per_block,tot_point_per_floor_point)
        equip_sum.to_excel(excel_writer, sheet_name='Equipment Summary', startrow= 2, startcol=1)
        create_equip_sum_format(excel_writer, 'Equipment Summary',equip_sum_row_num, equip_sum, equip_sum_count_list,block_list,equip_table_row_num,tot_point_per_block_equip, tot_point_per_floor_equip)
        # iterate each MCC create excel

        point_list,point_equip_type_set = get_points_list(point_file)
        df_equip = pd.read_csv(equip_file)
        grouped =  df_equip.groupby(['MCCLocation'])

        for MCCname, group in grouped:
            point_record_df,row_number_list, floor_name_list,total_BMS_point_list_per_floor = createPointsDF2(group, point_equip_type_set, point_list)
            create_point_list_sheet_format(point_record_df,MCCname[0],row_number_list, floor_name_list,total_BMS_point_list_per_floor,spare_points, excel_writer)
    return equip_sum, equip_sum_row_num, equip_sum_count_list, equip_table_row_num, block_list, point_sum,point_sum_row_num,point_sum_count_list,point_table_row_num, report_name

def copy_project_sum_win32(report_name):
    file_path = os.path.abspath('\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Template\\BMS Export Template.xlsx')
 
    path1 = file_path
    #xl=Dispatch("Excel.Application",pythoncom.CoInitialize())
    path2 = report_name
   


    wb_res = xl.Workbooks.Open(Filename=path1)
    wb_des = xl.Workbooks.Open(Filename=path2)

    ws1 = wb_res.Worksheets(1)
    ws1.Copy(Before=wb_des.Worksheets(1))
    

    wb_res.Close(SaveChanges=False)
    wb_des.Close(SaveChanges=True)
    xl.Quit()
   
def copy_project_sum(report_name):
    file_path = os.path.abspath('\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Template\\BMS Export Template.xlsx')
    # Load the source workbook
    source_wb = openpyxl.load_workbook(file_path)
    # Get the source worksheet
    source_sheet = source_wb['Project Summary']


    # load workbook and add a new sheet
    target_wb =load_workbook(report_name)
    target_sheet = target_wb.create_sheet('Project Summary', 0)

    # Copy the style of each cell in the source sheet to the target sheet
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)
        target_cell.value = source_cell.value
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

    # Copy column widths
    for col_letter in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

    # Copy row heights
    for row_number in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row_number].height = source_sheet.row_dimensions[row_number].height

    # Copy merged cells
    for cell_range in source_sheet.merged_cells:
        target_sheet.merge_cells(cell_range.coord)



    # Save the target workbook
    target_wb.save(report_name)   
    target_wb.close()




def edit_project_summary(projectName, current_date,report_name):
   
    wb = load_workbook(report_name)
    ws = wb['Project Summary']
    projectFile = Path(f'\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Points Schedule - {projectName} - {current_date.day:02d}-{current_date.month:02d}-{current_date.year}.csv\\Projects.csv')
    dfProject = pd.read_csv(projectFile)
    ws['C2'] = dfProject.loc[0]['JobName']
    ws['C4'] = dfProject.loc[0]['JobName']
    ws['C6'] = dfProject.loc[0]['Client']
    ws['C8'] = dfProject.loc[0]['Designer']
    ws['F6'] = dfProject.loc[0]['Verifier']
    ws['F8'] = dfProject.loc[0]['BuildingType']
    
    formatted_date = current_date.strftime('%d/%m/%Y')
    ws['F2'] = formatted_date


    # insert image
    img = Image('logo.png')
    ws.add_image(img, 'B2')

    wb.save(report_name)

def edit_electrical_loads(qty_floor_mcc_df,report_name):
    wb = load_workbook(report_name)
    ### add bottom border
    bottom_border = Border(left=Side(style='dashed'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    
    cell_border = Border(left=Side(style='dashed'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))
    left_cell_border = Border(left=Side(style='medium'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))

    right_cell_border = Border(left=Side(style='dashed'), 
                     right=Side(style='medium'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))
    left_corner_cell_border= Border(left=Side(style='medium'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    right_corner_cell_border= Border(left=Side(style='dashed'), 
                     right=Side(style='medium'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    bolder_cell = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    left_thick = Border(left = Side(style = "medium"))
    bottom_thick = Border(bottom = Side(style = "medium"))
    top_thick = Border(top = Side(style = "medium"))
    left_dash =  Border(left = Side(style = "dashed"))

    tot_font = Font(name='Cambria', bold=True,size = 9)
    alignment = Alignment( vertical='center')
    
    for key in qty_floor_mcc_df["MCC"].unique():
       
    ## change the table TITLE to MCC name
        wb.active = wb[f"{key} - Electrical Loads"]
        ws = wb.active
        ws['B2'].value = f"{key} - Electrical Loads"
#####  assistance function
        def copy_range(range_str, sheet, offset):

            """ Copy cell values and style to the new row using offset"""
            for row in rows_from_range(range_str):
                for cell in row:
                    if sheet[cell].value is not None:  # Don't copy other cells in merged unit
                        dst_cell = sheet[cell].offset(row=offset, column=0)
                        src_cell = sheet[cell]

                        ### Copy Cell value
                        dst_cell.value = src_cell.value

                        ### Copy Cell Styles
                        dst_cell.font = copy(src_cell.font)
                        dst_cell.alignment = copy(src_cell.alignment)
                        dst_cell.border = copy(src_cell.border)
                        dst_cell.fill = copy(src_cell.fill)
                        dst_cell.number_format = src_cell.number_format
   
        def get_merge_list(r_range, r_offset):
        
            """ Create a list of new cell merges from the existing row"""
            area = CellRange(r_range)  # Range to check for merged cells
            mlist = []  # List of merged cells on existing row offset to the new row
            for mc in ws.merged_cells:
                if mc.coord not in area:
                    continue
                cr = CellRange(mc.coord)
                cr.shift(row_shift=r_offset)
                mlist.append(cr.coord)
            return mlist
        
        # update value in each sheet
        qty_floor_mcc_df.sort_values(["MCC", "floor","equip_name","QTY"])
        equip_group_by_floor = qty_floor_mcc_df.groupby(["MCC","floor","equip_name"]).agg({"QTY":"sum"})
        start_row = 21
        
        floor_list = qty_floor_mcc_df[qty_floor_mcc_df['MCC'] == key]["floor"].unique()
        for num, f in enumerate(floor_list):  # return the floor only in that MCC
            # in each table in sheet. write equip 1 by 1
            # equi_qty_df = equip_group_by_floor.xs((key,f), level = ["MCC","floor"]).reset_index()["equip_name","QTY"]
            equi_qty_df = qty_floor_mcc_df[qty_floor_mcc_df['floor'] == f][["equip_name","QTY"]]
            # after filter, the index needs to reset into 0 in order to the iteration
            equi_qty_df = equi_qty_df.reset_index(drop = True)
            equip_num_floor = 0
            for  i in range(len(equi_qty_df)):
                
                # print(f"MCC: {key}. Floor: {f}")
                # print(equi_qty_df.loc[i,"equip_name"],equi_qty_df.loc[i,"QTY"])
                # repeat qty row for i equipment 
                
                for q in range(equi_qty_df.loc[i,"QTY"]):
                    equip_num_floor += 1
                    if equi_qty_df.loc[i,"equip_name"].count('-') > 3: # if it is "FCU-01-01to FCU-01-50" pattern
                        ws[f'B{start_row + q}'].value = equi_qty_df.loc[i,"equip_name"].split("-")[0]
                    else:
                        ws[f'B{start_row + q}'].value = equi_qty_df.loc[i,"equip_name"]
                    ws.column_dimensions['B'].width = 25


                    # the cell style
                    for col in range (2,28):
                        if col == 2 :
                            ws.cell(row = start_row + q, column = col ).border = left_cell_border
                        elif col == 27:
                            ws.cell(row = start_row + q, column = col ).border = right_cell_border
                        else:
                            ws.cell(row = start_row + q, column = col ).border = cell_border
            
       

               
                # finish one equipment iteration 
                start_row = start_row + q + 1
                # 8 is the line between 2 table + 3 for total equip num

            # finish one table in one floor  + num of empty line among 2 tables
            # last time no need to copy
            if num != len(floor_list) - 1:
                #### copy the title STYLE for all merged cell
                range_str = 'B15:AA20'
                #find header start row
                # -14 (startrow  start from 14) + 4 (4 empty row between 2 table)
                row_offset = start_row - 14 + 4
                
                ### Create a range list for merged cells on new row
                
                new_merge_list = get_merge_list(range_str, row_offset)
                
                # ### Create merged cells on new row
                for nm in new_merge_list:
                    ws.merge_cells(nm)

                ### Copy cell values to new row
                copy_range(range_str, ws, row_offset)

                # update header name
                ws[f'B{start_row + 5 }'].value = f"{key} - {f}"

                #### bottom border style
                ## bottom border is the previou row 
                for col in range(2,28):
                    if col == 3 or col == 5 or col == 6 or (col >=9 and col < 19) or (col >=21 and col <24):
                        for m in range(7,11):
                            ws.cell(row = start_row + m, column = col).border = left_dash


                    ws.cell(row = start_row + 10, column = col).border = bottom_thick # title lower border thick
                    ### set left border in title area
                    if col == 2 or  col == 4 or  col == 7 or  col == 8 or  col == 19 or  col >= 24 :
                        for c in range (7,10):
                            ws.cell(row = start_row + c, column = col).border = left_thick
                        ws.cell(row = start_row + 10, column = col).border = left_corner_cell_border
                    ws.cell(row = start_row , column = col).border =top_thick
                    ws.cell(row = start_row + 4, column = col).border = bottom_thick
                    ws.cell(row = start_row + 5, column = col).border = bolder_cell  # +5: header row
                    ws.cell(row = start_row + 6, column = col).border = bolder_cell   # +6: title row

                  
                ## last column
                for c in range(5,11):
                    ws.cell(row = start_row + c, column = 28).border = left_thick 
                   
         


           

            # write the total equip num
            ws.cell(row = start_row + 1, column = 2 ).border = bolder_cell
            ws.cell(row = start_row + 1, column = 2 ).font = tot_font
            ws.cell(row = start_row + 1, column = 2 ).alignment = alignment
            ws.row_dimensions[start_row + 1].height = 15
            ws.row_dimensions[start_row + 6].height = 30
            ws.row_dimensions[start_row + 5].height = 45
            ws.merge_cells(f"B{start_row + 1}:C{start_row + 1}")
            ws[f'B{start_row + 1}'].value = f"Total Equipment Number({f})"
            ws.cell(row = start_row + 1, column = 4 ).border = bolder_cell
            ws.cell(row = start_row + 1, column = 4 ).alignment = alignment
            ws.cell(row = start_row + 1, column = 4 ).font = tot_font
            ws[f'D{start_row + 1}'].value = equip_num_floor 
            
         
            
             
            start_row  += 11 
            

    ws.cell(row = start_row , column = 2 ).border = bolder_cell
    ws.cell(row = start_row, column = 2 ).font = tot_font
    ws.cell(row = start_row + 1, column = 2 ).alignment = alignment
    ws.merge_cells(f"B{start_row}:C{start_row}")
    ws[f'B{start_row }'].value = "Total Equipment Number:"
    ws.row_dimensions[start_row ].height = 15
    ws.cell(row = start_row , column = 4 ).border = bolder_cell
    ws.cell(row = start_row, column = 4 ).font = tot_font
    ws.cell(row = start_row, column = 4 ).alignment = alignment
    ws[f'D{start_row}'].value = qty_floor_mcc_df.groupby(["MCC"]).agg({"QTY":"sum"}).loc[key,"QTY"]
        

    final_row = start_row - 11
    #### add formula  
    for col in range(12, 16):
        col_letter = openpyxl.utils.get_column_letter(col)
        ## park load formula
        ws[f'{col_letter}7'].value = f'=SUM({col_letter}{21}:{col_letter}{final_row})' 
        # peak load formula
        ws[f'{col_letter}8'] = f'=IF($P$4="3Ph",(({col_letter}7*1000)/((SQRT(3))*400)),(({col_letter}7*1000)/230))' 
        # diversified peal load amps
        ws[f'{col_letter}10'] = f'={col_letter}7 * {col_letter}8'


    wb.save(report_name)



 







def copy_electrical_loads_title_win32(qty_per_MCC,report_name):

   #file_path = os.path.abspath('\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Template\\BMS Export Template.xlsx')
 
   # path1 = file_path
 
    path2 = report_name
    #xl = Dispatch("Excel.Application")


    wb_res = xl.Workbooks.Open(Filename=path1)
    wb_des = xl.Workbooks.Open(Filename=path2)

    ws_res = wb_res.Worksheets(2)
  

    for key, innerdict in qty_per_MCC.items():
        
            
    
        ws2 = wb_des.Worksheets.Add(After = wb_des.Worksheets['Project Summary'])
        ws2.Name = f"{key} - Electrical Loads"
        
        # Copy a range from the original worksheet
        # row of equip qty + offset
        start_row_sum = sum(innerdict.values()) + 12
        ws_res.Range("A1:AA10").Copy(ws2.Range("A1:AA10"))
        ws_res.Range("B43:AA49").Copy(ws2.Range(f"B{start_row_sum}:AA{start_row_sum + 6}"))


    wb_res.Close(SaveChanges=False)
    wb_des.Close(SaveChanges=True)
    xl.Quit()

def copy_electrical_loads_title(qty_floor_mcc_df,report_name):
    
    file_path = os.path.abspath('\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Template\\BMS Export Template.xlsx')
    # Load the source workbook
    source_wb = openpyxl.load_workbook(file_path)
    # Get the source worksheet
    source_sheet = source_wb['LG Floor - Electrical Loads']


    # load workbook and add a new sheet
    target_wb =load_workbook(report_name)
    
    for key in qty_floor_mcc_df["MCC"].unique():
        target_sheet = target_wb.create_sheet(f"{key} - Electrical Loads", 1)

        # Copy the style of each cell in the source sheet to the target sheet
        for (row, col), source_cell in source_sheet._cells.items():
            target_cell = target_sheet.cell(column=col, row=row)
            target_cell.value = source_cell.value
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        # Copy column widths
        for col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

        # Copy row heights
        for row_number in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row_number].height = source_sheet.row_dimensions[row_number].height

        # Copy merged cells
        for cell_range in source_sheet.merged_cells:
            target_sheet.merge_cells(cell_range.coord)



    # Save the target workbook
    target_wb.save(report_name)   
    target_wb.close()



def equip_summary_format(equip_sum,equip_sum_row_num,equip_sum_count_list,equip_table_row_num,block_list,report_name):
  
    path = report_name
    wb = load_workbook(path)
    ws = wb['Equipment Summary']
    bottom_border = Border(left=Side(style='dashed'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    
    top_thick_border = Border(top=Side(style='medium'), 
                    )
    
    cell_border = Border(left=Side(style='dashed'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))
    left_cell_border = Border(left=Side(style='medium'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))

    right_cell_border = Border(left=Side(style='dashed'), 
                     right=Side(style='medium'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))
    left_corner_cell_border= Border(left=Side(style='medium'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    right_corner_cell_border= Border(left=Side(style='dashed'), 
                     right=Side(style='medium'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    
    font = Font(name='Cambria',size = 8)
    left_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    tot_col = equip_sum.shape[1] + len(equip_sum.index[0])

    offset = 0

    for i, block in enumerate(block_list):
        
    ####### 1: 1ST TABLE FORMAT 
        for row in range(equip_sum_row_num[i] -2):
            for col in range(1, tot_col  + 1):
                # start from 3rd row, 2nd col
               
                start_row = row + 4 + offset
                start_col = col + 1
                ws.cell(row = start_row, column = start_col ).font = font
                ws.cell(row = start_row, column = start_col ).alignment = center_align
                #### start from 1
                if col == 1 or col == 2:

                    ws.cell(row = start_row, column = start_col ).border = left_cell_border
                
                elif (col == tot_col) or (col == len(equip_sum.index[0])):
                    ws.cell(row = start_row, column = start_col ).border = right_cell_border
                elif (col == 2) or (col == 3):
                    ws.cell(row = start_row, column = start_col ).alignment = left_align
                else:
                    ws.cell(row = start_row, column = start_col ).border = cell_border
      
        


        ####### 2: 2ND TABLE FORMAT 
    
       
        # -2 : 1 empty row, 1 2nd title row
        bd_row_num = equip_table_row_num[i] - equip_sum_row_num[i] - 2
        
        # start from 1st table row + 6, 2nd col
      
        start_bd_row = equip_sum_row_num[i]  + 4 + offset

     
        for row2 in range(bd_row_num):
            for col2 in range(1, tot_col  + 1):    
                start_row2 = start_bd_row + row2 
                start_col2 =  col2 + 1  
                ws.cell(row = start_row2, column = start_col2 ).font = font
                ws.cell(row = start_row2, column = start_col2 ).alignment = center_align
                ####There are 2 logic need to check in one loop
                if (col2 == 2) or (col2 == 3):
                    ws.cell(row = start_row2, column = start_col2 ).alignment = left_align
                if col2 == 1 or col2 == 2:
                    ws.cell(row = start_row2, column = start_col2 ).border = left_cell_border
                elif (col2 == tot_col) or (col2 == len(equip_sum.index[0])):
                    ws.cell(row = start_row2, column = start_col2 ).border = right_cell_border 
                else:
                    ws.cell(row = start_row2, column = start_col2 ).border = cell_border
        
        
        # 3: SET BOTTOM BORDER FOR 2 TABLES
        # border under each equipment in breakdown
        start = 3 + equip_sum_row_num[i] + offset
    
        for row in equip_sum_count_list[block]:
            for col in range(1, 10):
                    start_col = col + 1
                
                    ws.cell(row = start + row, column = start_col).border = bottom_border
                    ws.cell(row = equip_sum_row_num[i] + 1 + offset , column = start_col).border = bottom_border
                    ws.cell(row = equip_sum.shape[0] + 3 , column = start_col).border = bottom_border

                    if col == 1 or col == 2:
                                # 2nd table equip bottom border
                                ws.cell(row = start + row, column = start_col ).border = left_corner_cell_border
                                ## 1st table bottom border
                                ws.cell(row = equip_sum_row_num[i] + 1 + offset , column = start_col).border = left_corner_cell_border
                                ## 2nd table
                                ws.cell(row =equip_sum.shape[0] + 3, column = start_col ).border = left_corner_cell_border
                    elif (col == 9) or (col == 4) :
                                # 1st table bottom border
                                ws.cell(row = equip_sum_row_num[i] + 1 + offset, column = start_col).border = right_corner_cell_border
                                ws.cell(row = start + row, column = start_col ).border = right_corner_cell_border
                                ## 2nd table
                                ws.cell(row = equip_sum.shape[0] + 3, column = start_col ).border = right_corner_cell_border
            start = start + row 
                
        offset += equip_table_row_num[i] + 1

    wb.save(path)
    wb.close()


def point_summary_format(point_sum,point_sum_row_num,point_sum_count_list,point_table_row_num, block_list,floor_num_dict,report_name):
    path = report_name
    wb = load_workbook(path)
    ws = wb['Points Summary']
    bottom_border = Border(left=Side(style='dashed'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    
    cell_border = Border(left=Side(style='dashed'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))
    left_cell_border = Border(left=Side(style='medium'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))

    right_cell_border = Border(left=Side(style='dashed'), 
                     right=Side(style='medium'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))
    left_corner_cell_border= Border(left=Side(style='medium'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    right_corner_cell_border= Border(left=Side(style='dashed'), 
                     right=Side(style='medium'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='medium'))
    
    font = Font(name='Cambria',size = 8)
    left_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    tot_col = point_sum.shape[1] + len(point_sum.index[0])

    offset = 0
    for i, block in enumerate(block_list):

    
        ####### 1: 1ST TABLE FORMAT 
        for row in range(point_sum_row_num[i] - 2):
            for col in range(1, tot_col  + 1):
                # start from 3rd row, 2nd col
                start_row = row + 4 + offset
                start_col = col + 1
                ws.cell(row = start_row, column = start_col ).font = font
                ws.cell(row = start_row, column = start_col ).alignment = center_align
                #### start from 1
                if col == 1 or col == 2:
                    ws.cell(row = start_row, column = start_col ).border = left_cell_border
                elif (col == tot_col) or (col == len(point_sum.index[0])):
                    ws.cell(row = start_row, column = start_col ).border = right_cell_border             
                else:
                    ws.cell(row = start_row, column = start_col ).border = cell_border

                if (col == 2) or (col == 3):
                    ws.cell(row = start_row, column = start_col ).alignment = left_align
                if(ws.cell(row=start_row, column=3).value == "Floor Area"):
                 
                    ws[f'C{start_row}'] = f'Floor Area ({floor_num_dict[block]})'


        ####### 2: 2ND TABLE FORMAT 
        ### -2 because 1 empty row, 1 2nd table title
        bd_row_num = point_table_row_num[i] - point_sum_row_num[i] - 2
        
        # start from 1st table row + 1 emptyrow(begining of table) + 1 empty row(end of table) + 1: 2nd table title + 1(cell begining from 1 )
        start_bd_row =  point_sum_row_num[i] + 4 + offset
        # row start from 0. 
        for row in range(bd_row_num):
            for col in range(1, tot_col  + 1):    
                start_row = start_bd_row + row
                start_col =  col + 1  
                ws.cell(row = start_row, column = start_col ).font = font
                ws.cell(row = start_row, column = start_col ).alignment = center_align
                if (col == 2) or (col == 3):
                    ws.cell(row = start_row, column = start_col ).alignment = left_align 

                if col == 1 or col == 2:
                    ws.cell(row = start_row, column = start_col ).border = left_cell_border
                elif (col == tot_col) or (col == len(point_sum.index[0])):
                    ws.cell(row = start_row, column = start_col ).border = right_cell_border
                
                else:
                    ws.cell(row = start_row, column = start_col ).border = cell_border
        
        # 3: SET BOTTOM BORDER FOR 2 TABLES
        # border under each equipment in breakdown
        #4: 1 emptyrow(begining of table) + 1 empty row(end of table) + 1: 2nd table title 
        start = point_sum_row_num[i] + 3 + offset
        # row start from boundary
        for row in point_sum_count_list[block]:
            for col in range(2,9):
                   
                    ws.cell(row = start + row, column = col).border = bottom_border
                    ws.cell(row = point_sum_row_num[i] + 1 , column = col).border = bottom_border
                    ws.cell(row = point_sum.shape[0] + 3 , column = col).border = bottom_border

                    if col == 2 or col == 3:
                                # 2nd table equip bottom border
                                ws.cell(row = start + row, column = col ).border = left_corner_cell_border
                                ## 1st table bottom border
                                ws.cell(row = point_sum_row_num[i] + 1 + offset , column = col).border = left_corner_cell_border
                                ## 2nd table
                                ws.cell(row =point_sum.shape[0] + 3, column = col ).border = left_corner_cell_border
                    elif (col == 8) or (col == 5) :
                                ## 1st table bottom border
                                ws.cell(row = point_sum_row_num[i] + 1 + offset , column = col).border = right_corner_cell_border

                                ws.cell(row = start + row, column = col ).border = right_corner_cell_border
                                ## 2nd table
                                ws.cell(row = point_sum.shape[0] + 3, column = col ).border = right_corner_cell_border
            start = start + row 
        offset += point_table_row_num[i] + 1
        
            

    wb.save(path)
    wb.close()

def apply_border_again(equip_sum_row_num,equip_table_row_num,report_name):
    
    wb = load_workbook(report_name)
    ws = wb['Equipment Summary']

    top_thick_border = Border(top=Side(style='medium'))
    top_bottom_thick_border = Border(top=Side(style='medium'),bottom = Side(style='medium'))
    for col in [2,10]:
        ws.cell(row = equip_sum_row_num[0] + 2, column = col ).border = top_bottom_thick_border
        ws.cell(row = equip_table_row_num[0] + 2, column = col ).border = top_thick_border
        
    wb.save(report_name)
    wb.close()


def get_final_report(projectName):

    current_date = date.today()
    #current_date = datetime.datetime(2023, 11, 15)
    equip_sum, equip_sum_row_num, equip_sum_count_list, equip_table_row_num, block_list,point_sum, point_sum_row_num,point_sum_count_list, point_table_row_num,report_name = createReportExcel(
                      f'\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Points Schedule - {projectName} - {current_date.day:02d}-{current_date.month:02d}-{current_date.year}.csv\\Points.csv',
                      f'\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Points Schedule - {projectName} - {current_date.day:02d}-{current_date.month:02d}-{current_date.year}.csv\\Equipment.csv',
                      f'\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Points Schedule - {projectName} - {current_date.day:02d}-{current_date.month:02d}-{current_date.year}.csv\\Projects.csv')
   
    #openpyxl to format the cell font
    ## need get equip per MCC {'MMC -A' : {"fcu" : 6, "ahu":9}}
    ### num of equipment = row of that equip
    _, qty_floor_mcc_df ,_,floor_num_dict= get_equipments_list(f'\\\\eeazurefilesne.file.core.windows.net\\generalshare\\Ethos Digital\\BMS Points Generator Reports\\Points Schedule - {projectName} - {current_date.day:02d}-{current_date.month:02d}-{current_date.year}.csv\\Equipment.csv') 
    point_summary_format(point_sum,point_sum_row_num,point_sum_count_list,point_table_row_num, block_list,floor_num_dict,report_name)
    equip_summary_format(equip_sum,equip_sum_row_num,equip_sum_count_list,equip_table_row_num, block_list,report_name)
   
    
    
    copy_project_sum(report_name)
    # ##### EDIT copied sheet(Project summary and electrical loads)  
    edit_project_summary(projectName,current_date,report_name)


    if not qty_floor_mcc_df.empty:
        copy_electrical_loads_title(qty_floor_mcc_df,report_name)
        edit_electrical_loads(qty_floor_mcc_df,report_name)
    
    #openpyxl didn't preserve all the original information when it loaded the workbook, some of the original formatting might be lost, so we need to reapply missing feature
    ## in our case,. equipment summary 2 bottom border has issue
    apply_border_again(equip_sum_row_num,equip_table_row_num,report_name)

if __name__=="__main__":
    #get_final_report('Camden Yard Script Test')
    get_final_report('Linkedin Test')